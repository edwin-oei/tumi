import pymysql
import datetime
import time
import os
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

import connection
from connection import db_ip_address
from connection import db_user
from connection import db_password
from connection import db_name
from connection import student_data
from connection import event_folder
from connection import receipt_temp_folder
from connection import receipt_perm_folder
from connection import email_sender
from connection import email_sender_password
from connection import email_hiwi
from connection import server_out
from connection import port_out

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders



def send_KickEmail(fnameWL, lnameWL, emailWL, subject, eventDate):
    global email_user
    global email_db_password
    global server_out
    global port_out

    email_send = emailWL

    year, month, day = map(str, eventDate.split('-'))
    eventDate = day + "/" + month + "/" + year

    subject2 = "TUMi: " + subject + " You Are Relegated to the Waiting List"

    msg = MIMEMultipart()
    msg['From'] = email_user
    msg['To'] = email_send
    msg['Subject'] = subject2

    body = "Hi " + fnameWL + " " + lnameWL + ",\n\nDue to organisational reasons, we regret to inform you that you have just been relegated to the waiting list for the event: " + subject + " on " + eventDate + ".\nDeregistration is 2 working days before the event.\n\n\nRegards,\n\nThomas Bergmann, M.A.\nOffice for Incoming Exchange Students / TUMi\n\nTechnical University of Munich\nTUM International Centre\n\nArcisstr. 21\n80333 Munich\n\nTel: + 49 89 289 25477"
    msg.attach(MIMEText(body, 'plain'))

    #msg.attach(part)

    # msg.attach(part)
    text = msg.as_string()
    server = smtplib.SMTP(server_out, port_out)
    server.starttls()
    server.login(email_user, email_db_password)
    server.sendmail(email_user, email_send, text)
    print("\nEmail kicked to the waiting list sent to:", email_send)
    server.quit()



def sendShiftReceipt(fnameWL, lnameWL, emailWL, subject, eventDate):
    global email_user
    global email_db_password
    global server_out
    global port_out

    email_send = emailWL

    year, month, day = map(str, eventDate.split('-'))
    eventDate = day + "/" + month + "/" + year

    subject2 = "TUMi: " + subject + " You Are Off the Waiting List"

    msg = MIMEMultipart()
    msg['From'] = email_user
    msg['To'] = email_send
    msg['Subject'] = subject2

    body = "Hi " + fnameWL + " " + lnameWL + ",\n\nWe would like to inform you that you are now off the waiting list and in the official list of participants for the event: " + subject + " on " + eventDate + ".\nDeregistration is 2 working days before the event.\n\n\nRegards,\n\nThomas Bergmann, M.A.\nOffice for Incoming Exchange Students / TUMi\n\nTechnical University of Munich\nTUM International Centre\n\nArcisstr. 21\n80333 Munich\n\nTel: + 49 89 289 25477"
    msg.attach(MIMEText(body, 'plain'))

    #msg.attach(part)

    # msg.attach(part)
    text = msg.as_string()
    server = smtplib.SMTP(server_out, port_out)
    server.starttls()
    server.login(email_user, email_db_password)
    server.sendmail(email_user, email_send, text)
    print("\nEmail off the waiting list sent to:", email_send)
    server.quit()

def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


def wrongEntry():
    print("Wrong entry. Please try again.\n")
    #time.sleep(2)


def jeinError():
    print("Please enter either 1 or 0.\n")
    #time.sleep(1.5)


def confirm_Kick(select, db, cursor):

    sql = "SELECT * FROM events WHERE eventCode = '%s'" % (select)
    cursor.execute(sql)
    result2 = cursor.fetchall()
    for column in result2:
        date = column[0]
        code = column[1]
        maxPart = column[3]
        maxWait = column[5]
        price = column[6]
        desc = column[7]
    print("Date\t\t\t  Code\t\t\tMax Participants\t\tMax Waiting List\t\tPrice\t\tDescription")
    print(date, "\t\t ", code, "\t\t\t", maxPart, "\t\t\t\t\t", maxWait, "\t\t\t\t   ", price, "\t\t  ", desc)

    i = 0
    while i == 0:
        jein = input("\nKick people from the regular list to the waiting list and confirm changes? (1) Yes\t(0) No")
        if jein == '1':
            i = 1
            db.commit()
            db.close()
        elif jein == '0':
            db.rollback()
            i = 1
            db.close()
        else:
            jeinError()

    if jein == '1':
        return 1
    elif jein == '0':
        return 0


def confirm_new(select, db, cursor):

    sql = "SELECT * FROM events WHERE eventCode = '%s'" % (select)
    cursor.execute(sql)
    result2 = cursor.fetchall()
    for column in result2:
        date = column[0]
        code = column[1]
        maxPart = column[3]
        maxWait = column[5]
        price = column[6]
        desc = column[7]
    print("Date\t\t\t  Code\t\t\tMax Participants\t\tMax Waiting List\t\tPrice\t\tDescription")
    print(date, "\t\t ", code, "\t\t\t", maxPart, "\t\t\t\t\t", maxWait, "\t\t\t\t   ", price, "\t\t  ", desc)

    i = 0
    while i == 0:
        jein = input("Confirm changes? (1) Yes\t(0) No")
        if jein == '1':
            i = 1
            db.commit()
            db.close()
        elif jein == '0':
            db.rollback()
            i = 1
            db.close()
        else:
            jeinError()

    if jein == '1':
        return 1
    elif jein == '0':
        return 0


def confirm(select, sql1, db, cursor):

    cursor.execute(sql1)
    sql2 = "SELECT * FROM events WHERE eventCode = '%s'" % (select)
    cursor.execute(sql2)
    result2 = cursor.fetchall()
    for column in result2:
        date = column[0]
        code = column[1]
        maxPart = column[3]
        maxWait = column[5]
        price = column[6]
        desc = column[7]
    print("Date\t\t\t  Code\t\t\tMax Participants\t\tMax Waiting List\t\tPrice\t\tDescription")
    print(date, "\t\t ", code, "\t\t\t", maxPart, "\t\t\t\t\t\t", maxWait, "\t\t\t\t   ", price, "\t\t  ", desc)

    i = 0
    while i == 0:
        jein = input("Confirm changes? (1) Yes\t(0) No")
        if jein == '1':
            i = 1
            db.commit()
            db.close()
        elif jein == '0':
            db.rollback()
            i = 1
            db.close()
        else:
            jeinError()


def newExcel(eventDate, eventCode, excelPrice, eventDesc, maxPart):
    global event_folder

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"
    wb.create_sheet("Waiting_List")
    ws2 = wb["Waiting_List"]
    my_cell = ws['A1']
    my_cell.value = eventDesc

    fontTitle = Font(b=True, size=15, name='Arial')
    al = Alignment(horizontal="center", vertical="center")
    style_range(ws, 'A1:H2', font=fontTitle, alignment=al)

    myCell2 = ws['A3']
    eventDate = str(eventDate)
    year, month, day = map(str, eventDate.split('-'))
    eventDate = day+"/"+month+"/"+year
    eventDate = "Date: " + eventDate
    myCell2.value = eventDate
    fontDate = Font(size=13, name='Arial')
    alDate = Alignment(horizontal="left", vertical="center")
    style_range(ws, 'A3:H3', font=fontDate, alignment=alDate)

    myCell3 = ws['A4']
    excelPrice = "Fee:  €" + excelPrice
    myCell3.value = excelPrice
    fontDate = Font(size=13, name='Arial')
    alDate = Alignment(horizontal="left", vertical="center")
    style_range(ws, 'A4:H4', font=fontDate, alignment=alDate)

    ws['A5'] = 'Nr.'
    ws['B5'] = 'First Name'
    ws['C5'] = 'Last Name'
    ws['D5'] = 'Email'
    ws['E5'] = 'Country'
    ws['F5'] = 'Department'
    ws['G5'] = 'E / D'
    ws['H5'] = 'Signature'

    for i in range(1, 9):
        ws.cell(row=5, column=i).font = Font(name='Arial', size=11, bold=True)
        ws.cell(row=5, column=i).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=5, column=i).fill = PatternFill("solid", fgColor="B9D3EE")
        ws.cell(row=5, column=i).border = Border(top=Side(border_style='medium', color='000000'),
                                                 left=Side(border_style='medium', color='000000'),
                                                 bottom=Side(border_style='medium', color='000000'),
                                                 right=Side(border_style='medium', color='000000'))

    ws2['A1'] = 'First Name'
    ws2['B1'] = 'Last Name'
    ws2['C1'] = 'Email'
    ws2['D1'] = 'Country'
    ws2['E1'] = 'Department'
    ws2['F1'] = 'E / D'


    for i in range(1, 7):
        ws2.cell(row=1, column=i).font = Font(name='Arial', size=11, bold=True)
        ws2.cell(row=1, column=i).alignment = Alignment(horizontal='center', vertical='center')
        ws2.cell(row=1, column=i).fill = PatternFill("solid", fgColor="B9D3EE")
        ws2.cell(row=1, column=i).border = Border(top=Side(border_style='medium', color='000000'),
                                                 left=Side(border_style='medium', color='000000'),
                                                 bottom=Side(border_style='medium', color='000000'),
                                                 right=Side(border_style='medium', color='000000'))

    eventCode = eventCode + ".xlsx"
    filePath = event_folder + eventCode
    wb.save(filePath)


def deleteExcel(eventCode):
    global event_folder

    excelName = eventCode + ".xlsx"
    filePath = event_folder + excelName
    os.remove(filePath)
    print("Corresponding excel file deleted.")
    #time.sleep(1.5)


def renameExcel(oldCode, newCode):
    global event_folder

    oldExcelName = oldCode + ".xlsx"
    newExcelName = newCode + ".xlsx"

    filePathOld = event_folder + oldExcelName
    filePathNew = event_folder + newExcelName
    os.rename(filePathOld, filePathNew)

    print("Corresponding excel file renamed to:", newExcelName)
    #time.sleep(2)


def changeDescExcel(newDesc, select):
    global event_folder

    fileName = select + ".xlsx"
    filePath = event_folder + fileName
    wb = load_workbook(filePath)
    ws = wb.active
    ws['A1'] = newDesc
    wb.save(filePath)


def changeDateExcel(newDate, select):
    global event_folder

    fileName = select + ".xlsx"
    filePath = event_folder + fileName
    wb = load_workbook(filePath)
    ws = wb.active
    newDate = str(newDate)
    year, month, day = map(str, newDate.split('-'))
    newDate = day + "/" + month + "/" + year
    newDate = "Date: " + newDate
    ws['A3'] = newDate
    wb.save(filePath)


def changePriceExcel(newExcelPrice, select):
    global event_folder

    fileName = select + ".xlsx"
    filePath = event_folder + fileName
    wb = load_workbook(filePath)
    ws = wb.active
    ws['A4'] = "Fee:  €" + newExcelPrice
    wb.save(filePath)


def newEvent():
    db = pymysql.connect(host=db_ip_address, user=db_user, passwd=db_password, db=db_name)
    cursor = db.cursor()

    eventCode = input("What is the code of the event? (Format: XXX_1)")

    k = 0
    while k == 0:
        sql1 = 'SELECT * FROM events WHERE eventCode = "%s"' % (eventCode)
        cursor.execute(sql1)
        result = cursor.fetchall()
        for column in result:
            exist = column[1]

        try:
            if exist == eventCode:
                print("Event code has been taken. Please choose another code.\n")
                #time.sleep(2)
                del exist
                eventCode = input("Enter another event code: (Format: XXX_1)")
        except:
            k = 1
            continue

    i = 0
    while i == 0:
        try:
            print("")
            date_entry = input("Date of the event in DD/MM/YYYY format:")
            day, month, year = map(int, date_entry.split('/'))
            eventDate = datetime.date(year, month, day)
            i = 1
        except:
            print("Please enter the date in the correct format.\n")
            #time.sleep(2)
            continue

    i = 0
    while i == 0:
        try:
            maxPart = int(input("What is the max number of participants?"))
            i = 1
        except:
            print("Please enter a number without any decimal places.")
            #time.sleep(2)

    i = 0
    while i == 0:
        try:
            maxWait = int(input("How many spots for the waiting list?"))
            i = 1
        except:
            print("Please enter a number without any decimal places.")
            #time.sleep(2)

    i = 0
    while i == 0:
        try:
            eventPrice = float(input("What is the price of the event?"))
            excelPrice = '%.2f' % eventPrice
            excelPrice = str(excelPrice)
            i = 1
        except:
            print("Please enter a number with max 2 decimal places.")
            #time.sleep(2)

    eventDesc = input("Description of the event?")

    sql5 = 'INSERT INTO events value("%s", "%s", 0, "%i", 0, "%i", "%.2f", "%s")' % (
        eventDate, eventCode, maxPart, maxWait, eventPrice, eventDesc)
    cursor.execute(sql5)

    print("Date\t\t\t  Code\t\t\tMax Participants\t\tMax Waiting List\t\tPrice\t\tDescription")
    print(eventDate, "\t\t ", eventCode, "\t\t\t", maxPart, "\t\t\t\t\t", maxWait, "\t\t\t\t   ", eventPrice, "\t\t  ", eventDesc)

    i = 0
    while i == 0:
        jein = input("Create the event? (1) Yes\t(0) No")
        if jein == '1':
            newExcel(eventDate, eventCode, excelPrice, eventDesc, maxPart)
            db.commit()
            db.close()
            i = 1
        elif jein == '0':
            db.rollback()
            db.close()
            i = 1
        else:
            jeinError()



def editEvent():
    db = pymysql.connect(host=db_ip_address, user=db_user, passwd=db_password, db=db_name)
    cursor = db.cursor()

    select = input("What is the code of the event to be edited?")
    sql = "SELECT * FROM events WHERE eventCode = '%s'" % (select)
    cursor.execute(sql)

    try:
        result = cursor.fetchall()
        for column in result:
            eventDate = column[0]
            eventCode = column[1]
            numPart = column[2]
            maxPart = column[3]
            waitPart = column[4]
            maxWait = column[5]
            eventPrice = column[6]
            eventDesc = column[7]

        print("\nDate\t\t\t  Code\t\t\tMax places\t\tMax waiting list\t\tPrice\t\tDescription")
        print(eventDate, "\t\t ", eventCode, "\t\t\t", maxPart, "\t\t\t\t\t", maxWait, "\t\t\t   ", eventPrice, "\t\t",
              eventDesc)

        j = 0
        while j == 0:
            editChoice = input("\nWhat would you like to edit?\n(1) Date\t\t\t(3) Max places\t\t\t(5) Price\n(2) Code\t\t\t(4) Max waiting list\t(6) Description")

            if editChoice == '1':
                i = 0
                while i == 0:
                    try:
                        date_entry = input('New date in in DD/MM/YYYY format')
                        day, month, year = map(int, date_entry.split('/'))
                        newDate = datetime.date(year, month, day)
                        sql1 = "UPDATE events SET date = '%s' WHERE eventCode = '%s'" % (newDate, select)
                        i = 1
                        j = 1
                        confirm(select, sql1, db, cursor)
                        changeDateExcel(newDate, select)
                    except:
                        print("Please enter a valid date in the required format.\n")
                        #time.sleep(2)
                        continue

            elif editChoice == '2':
                newCode = input("New code for the event: ")
                sql1 = "UPDATE events SET eventCode = '%s' WHERE eventCode = '%s'" % (newCode, select)
                j = 1
                confirm(newCode, sql1, db, cursor)
                renameExcel(select, newCode)

            elif editChoice == '3':
                i = 0
                while i == 0:
                    try:
                        confirm_jein = 0
                        newMax = int(input("New number of max spots: "))
                        j = 1   # Integer entered
                        #sql1 = "UPDATE events SET maxParticipants = '%s' WHERE eventCode = '%s'" % (newMax, select)
                        i = 1

                        if newMax >= maxPart:    # Case A: increase the max number of participants
                            if numPart <= newMax and waitPart == 0: # Case 1
                                sql1 = "UPDATE events SET maxParticipants = '%s' WHERE eventCode = '%s'" % (newMax, select)
                                confirm(select, sql1, db, cursor)
                            elif (newMax - numPart) >= waitPart and waitPart > 0:   # Case 2a
                                # everyone in waiting list to part list
                                excelFile = select + '.xlsx'
                                filePath = event_folder + excelFile
                                wb = load_workbook(filePath)
                                ws = wb['Sheet']
                                ws2 = wb['Waiting_List']

                                # databank update
                                sql1 = "UPDATE events SET maxParticipants = '%i' WHERE eventCode = '%s'" % (newMax, select)
                                sql2 = "UPDATE events SET numberOfParticipants = numberOfParticipants + '%i' WHERE eventCode = '%s'" % (waitPart, select)
                                sql3 = "UPDATE events SET inWaitingList = inWaitingList - '%i' WHERE eventCode = '%s'" % (waitPart, select)
                                cursor.execute(sql1)
                                cursor.execute(sql2)
                                cursor.execute(sql3)
                                confirm_jein = confirm_new(select, db, cursor)
                                if confirm_jein == 1:
                                    for k in range(newMax-numPart):
                                        fnameWL = ws2.cell(row=2, column=1).value
                                        lnameWL = ws2.cell(row=2, column=2).value
                                        emailWL = ws2.cell(row=2, column=3).value
                                        countryWL = ws2.cell(row=2, column=4).value
                                        departmentWL = ws2.cell(row=2, column=5).value
                                        edWL = ws2.cell(row=2, column=6).value
                                        sendShiftReceipt(fnameWL, lnameWL, emailWL, eventDesc, str(eventDate))  # inform via email
                                        ws.append(["", fnameWL, lnameWL, emailWL, countryWL, departmentWL, edWL])
                                        ws2.delete_rows(2)
                                    wb.save(filePath)
                                else:
                                    continue


                            elif (newMax - numPart) < waitPart and waitPart > 0:    # Case 2b
                                # everyone in waiting list to part list
                                excelFile = select + '.xlsx'
                                filePath = event_folder + excelFile
                                wb = load_workbook(filePath)
                                ws = wb['Sheet']
                                ws2 = wb['Waiting_List']

                                # databank update
                                sql1 = "UPDATE events SET maxParticipants = '%i' WHERE eventCode = '%s'" % (newMax, select)
                                sql2 = "UPDATE events SET numberOfParticipants = numberOfParticipants + '%i' WHERE eventCode = '%s'" % ((newMax - numPart), select)
                                sql3 = "UPDATE events SET inWaitingList = inWaitingList - '%i' WHERE eventCode = '%s'" % ((newMax - numPart), select)
                                cursor.execute(sql1)
                                cursor.execute(sql2)
                                cursor.execute(sql3)
                                confirm_jein = confirm_new(select, db, cursor)
                                if confirm_jein == 1:
                                    for k in range(newMax-numPart):
                                        fnameWL = ws2.cell(row=2, column=1).value
                                        lnameWL = ws2.cell(row=2, column=2).value
                                        emailWL = ws2.cell(row=2, column=3).value
                                        countryWL = ws2.cell(row=2, column=4).value
                                        departmentWL = ws2.cell(row=2, column=5).value
                                        edWL = ws2.cell(row=2, column=6).value
                                        sendShiftReceipt(fnameWL, lnameWL, emailWL, eventDesc, str(eventDate))  # inform via email
                                        ws.append(["", fnameWL, lnameWL, emailWL, countryWL, departmentWL, edWL])
                                        ws2.delete_rows(2)
                                    wb.save(filePath)
                                else:
                                    continue



                        elif newMax < maxPart:  # Case B: decrease the max number of participants
                            if numPart <= newMax:
                                confirm(select, sql1, db, cursor)
                            elif numPart > newMax:
                                # extend WL by (numParts - newMax)
                                if (numPart - newMax) > (maxWait - waitPart):
                                    sql1 = "UPDATE events SET maxParticipants = '%i' WHERE eventCode = '%s'" % (newMax, select)
                                    sql2 = "UPDATE events SET maxWaitingList = maxWaitingList + '%i' WHERE eventCode = '%s'" % (numPart - newMax - (maxWait - waitPart), select)
                                    sql3 = "UPDATE events SET inWaitingList = maxWaitingList WHERE eventCode = '%s'" % (select)
                                    sql4 = "UPDATE events SET numberOfParticipants = numberOfParticipants - '%i' WHERE eventCode = '%s'" % (numPart - newMax, select)
                                    cursor.execute(sql1)
                                    cursor.execute(sql2)
                                    cursor.execute(sql3)
                                    cursor.execute(sql4)
                                    confirm_jein = confirm_Kick(select, db, cursor)
                                else:
                                    sql1 = "UPDATE events SET maxParticipants = '%i' WHERE eventCode = '%s'" % (newMax, select)
                                    sql2 = "UPDATE events SET inWaitingList = inWaitingList + '%i' WHERE eventCode = '%s'" % ((numPart - newMax), select)
                                    sql3 = "UPDATE events SET numberOfParticipants = numberOfParticipants - '%i' WHERE eventCode = '%s'" % ((numPart - newMax), select)
                                    cursor.execute(sql1)
                                    cursor.execute(sql2)
                                    cursor.execute(sql3)
                                    print("pos1")
                                    confirm_jein = confirm_Kick(select, db, cursor)

                                if confirm_jein == 1:
                                    # put the unlucky ones in the first rows of WL
                                    num_of_unlucky = numPart - newMax
                                    excelFile = select + '.xlsx'
                                    filePath = event_folder + excelFile
                                    wb = load_workbook(filePath)
                                    ws = wb['Sheet']
                                    ws2 = wb['Waiting_List']

                                    for k in range(num_of_unlucky):
                                        fnameWL = ws.cell(row=5+numPart-k, column=2).value
                                        lnameWL = ws.cell(row=5+numPart-k, column=3).value
                                        emailWL = ws.cell(row=5+numPart-k, column=4).value
                                        countryWL = ws.cell(row=5+numPart-k, column=5).value
                                        departmentWL = ws.cell(row=5+numPart-k, column=6).value
                                        edWL = ws.cell(row=5+numPart-k, column=7).value
                                        send_KickEmail(fnameWL, lnameWL, emailWL, eventDesc, str(eventDate))  # inform via email
                                        ws2.insert_rows(2)
                                        total_column = 6
                                        column_list = [fnameWL, lnameWL, emailWL, countryWL, departmentWL, edWL]
                                        for column_counter in range(total_column):
                                            ws2.cell(row=2, column=column_counter+1).value = column_list[column_counter]
                                        ws.delete_rows(5+numPart-k)
                                    wb.save(filePath)
                                else:
                                    continue

                    except:
                        print("Please enter an integer.\n")
                        #time.sleep(2)

            elif editChoice == '4':
                i = 0
                while i == 0:
                    try:
                        new_maxWait = int(input("New max waiting list spots: "))
                        if new_maxWait < (waitPart):
                            print("\nPlease enter an integer bigger than", waitPart, "as there are already", waitPart, "people in the waiting list.\n")
                            #time.sleep(4)
                        else:
                            sql1 = "UPDATE events SET maxWaitingList = '%s' WHERE eventCode = '%s'" % (
                            new_maxWait, select)
                            i = 1
                            j = 1
                            confirm(select, sql1, db, cursor)

                    except:
                        print("Please enter an integer.\n")
                        #time.sleep(2)


            elif editChoice == '5':
                i = 0
                while i == 0:
                    try:
                        newPrice = float(input("New price for the event: "))
                        newExcelPrice = '%.2f' % newPrice
                        newExcelPrice = str(newExcelPrice)
                        sql1 = "UPDATE events SET price = '%s' WHERE eventCode = '%s'" % (newPrice, select)
                        changePriceExcel(newExcelPrice, select)
                        confirm(select, sql1, db, cursor)
                        i = 1
                        j = 1
                    except:
                        print("Please enter a number with max 2 decimal places.\n")
                        #time.sleep(2)

            elif editChoice == '6':
                newDesc = input("New description: ")
                sql1 = "UPDATE events SET description = '%s' WHERE eventCode = '%s'" % (newDesc, select)
                changeDescExcel(newDesc, select)
                confirm(select, sql1, db, cursor)
                j = 1

            else:
                wrongEntry()

    except:
        print("\n\n\n\n\nEvent does not exist.\n")
        #time.sleep(2)


def checkEvent():
    db = pymysql.connect(host=db_ip_address, user=db_user, passwd=db_password, db=db_name)
    cursor = db.cursor()

    sql = "SELECT date, eventCode, numberOfParticipants, maxParticipants, inWaitingList, maxWaitingList, price, description FROM events order by date ASC"
    cursor.execute(sql)
    db.commit()

    result = cursor.fetchall()
    print("\nDate\t\t\tCode\t\tPlaces left\t\tWaiting list spots left\t\tPrice\t\tDescription")
    for column in result:
        eventDate = column[0]
        eventCode = column[1]
        eventSpots = column[3] - column[2]
        waitSpots = column[5] - column[4]
        eventPrice = column[6]
        eventDescription = column[7]
        print(eventDate, "\t\t", eventCode, "\t\t   ", eventSpots, "\t\t\t\t\t", waitSpots, "\t\t\t\t\t", eventPrice,
              "\t\t  ", eventDescription)

    db.close()


def removeEvent():
    db = pymysql.connect(host=db_ip_address, user=db_user, passwd=db_password, db=db_name)
    cursor = db.cursor()

    select = input("What is the code of the event to be removed?")
    sql = "SELECT * FROM events WHERE eventCode = '%s'" % (select)
    cursor.execute(sql)

    try:
        result = cursor.fetchall()
        for column in result:
            eventDate = column[0]
            eventCode = column[1]
            numPart = column[2]
            inWait = column[4]
            eventPrice = column[6]
            eventDescription = column[7]

        print("\nDate\t\t\t  Code\t\t\tRegistered Participants\t\tIn Waiting List\t\tPrice\t\tDescription")
        print(eventDate, "\t\t ", eventCode, "\t\t\t\t", numPart, "\t\t\t\t\t\t", inWait, "\t\t\t   ", eventPrice, "\t\t",
              eventDescription)
        i = 0
        while i == 0:
            jein = input("Delete the event? (1) Yes\t(0) No")
            if jein == '1':
                sql1 = "DELETE FROM events WHERE eventCode = '%s'" % (select)
                deleteExcel(select)
                cursor.execute(sql1)
                db.commit()
                db.close()
                i = 1
            elif jein == '0':
                db.rollback()
                db.close()
                i = 1
            else:
                jeinError()

    except:
        print("\n\n\n\n\nEvent does not exist.\n")
        #time.sleep(1.5)


# main body

print("\n--------------Welcome to the event management program--------------\n")

i = 0
while i == 0:
    choice = input(
        "\nWhat would you like to do?\n(1) Add new event\t\t(3) Check event\n(2) Edit event\t\t\t(4) Remove event")

    if choice == '1':
        newEvent()
    elif choice == '2':
        editEvent()
    elif choice == '3':
        checkEvent()
    elif choice == '4':
        removeEvent()
    else:
        wrongEntry()