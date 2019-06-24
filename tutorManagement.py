import pymysql
import datetime
import time
import os
import pandas as pd
from pandas import ExcelWriter
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

import connection
from connection import db_ip_address
from connection import db_user
from connection import db_password
from connection import db_name
from connection import tutor_data

db = pymysql.connect(host=db_ip_address, user=db_user, passwd=db_password, db=db_name)
cursor = db.cursor()

sheet_name = 'Sheet1'

def jeinError():
    print("Please enter either 1 or 0.\n")



def wrongEntry():
    print("Wrong entry. Please try again.\n")



def update_Master_List(fname, lname, email):
    global tutor_data
    global sheet_name
    wb = load_workbook(tutor_data)
    ws = wb[sheet_name]
    ws.append([fname, lname, email])
    wb.save(tutor_data)
    wb.close()




def updatedEntry(result):

    for column in result:
        fname = column[0]
        lname = column[1]
        email = column[2]
    print(fname, lname, email)
    i = 0
    while i == 0:
        jein = input("Confirm entry (1) Yes\t(0) No")
        if jein == '1':
            db.commit()
            i = 1
            return 1
        elif jein == '0':
            db.rollback()
            i = 1
        else:
            jeinError()




def newFname(email, ws):

    new_fname = input("New first name:")
    sql = 'UPDATE tutor_data SET First_Name = "%s" WHERE Email = "%s"' % (new_fname, email)
    cursor.execute(sql)
    sql1 = "SELECT * FROM tutor_data WHERE Email = '%s'" % (email)
    cursor.execute(sql1)
    result = cursor.fetchall()
    if updatedEntry(result) == 1:    
        for rowNum in range(ws.max_row, 1, -1):
            if ws.cell(row=rowNum, column=3).value == email:
                ws.cell(row=rowNum, column=1).value = new_fname




def newLname(email, ws):

    new_lname = input("New last name:")
    sql = 'UPDATE tutor_data SET Last_Name = "%s" WHERE Email = "%s"' % (new_lname, email)
    cursor.execute(sql)
    sql1 = "SELECT * FROM tutor_data WHERE Email = '%s'" % (email)
    cursor.execute(sql1)
    result = cursor.fetchall()
    if updatedEntry(result) == 1:
        for rowNum in range(ws.max_row, 1, -1):
            if ws.cell(row=rowNum, column=3).value == email:
                ws.cell(row=rowNum, column=2).value = new_lname

    

def newEmail(email, ws):

    new_email = input("New email address:")
    sql = "UPDATE tutor_data SET Email = '%s' WHERE Email = '%s'" % (new_email, email)
    cursor.execute(sql)
    sql1 = "SELECT * FROM tutor_data WHERE Email = '%s'" % (new_email)
    cursor.execute(sql1)
    result = cursor.fetchall()
    if updatedEntry(result) == 1:
        for rowNum in range(ws.max_row, 1, -1):
            if ws.cell(row=rowNum, column=3).value == email:
                ws.cell(row=rowNum, column=3).value = new_email



def delete_Everyone(ws):

    sql = "DELETE FROM tutor_data"
    cursor.execute(sql)

    i = 0
    while i == 0:
        jein = input("Delete everyone. Are you sure? (1) Yes\t(0) No")
        if jein == '1':
            j = 0
            while j == 0:
                    jein2 = input("ARE YOU FUCKING SURE??? (1) Yes\t(0) No")
                    if jein2 == '1':
                        db.commit()
                        ws.delete_rows(2, ws.max_row)
                        i = 1
                        j = 1
                    elif jein2 == '0':
                        db.rollback()
                        i = 1
                        j = 1
                    else:
                        jeinError()
        elif jein == '0':
            db.rollback()
            i = 1
        else:
            jeinError()



def delete_Individual(email, ws):

    sql = "SELECT * FROM tutor_data WHERE Email = '%s'" % (email)
    cursor.execute(sql)
    result = cursor.fetchall()
    for column in result:
        fname = column[0]
        lname = column[1]

    try:
        print("\nData of", fname, lname, "will be deleted.")

        i = 0
        while i == 0:
            jein = input("Confirm? (1) Yes\t(0) No")
            if jein == '1':
                i = 1
                sql1 = "DELETE FROM tutor_data WHERE Email = '%s'" % (email)
                cursor.execute(sql1)
                db.commit()
                for rowNum in range(ws.max_row, 1, -1):
                    if ws.cell(row=rowNum, column=3).value == email:
                        ws.delete_rows(rowNum)
            elif jein == '0':
                db.rollback()
                i = 1
            else:
                jeinError()
    except:
        print("Person does not exist in the database to begin with.\n")



def delete_Mode():

    print("\n-------------Delete mode-------------\n")
    
    wb = load_workbook(tutor_data)
    ws = wb[sheet_name]
    i = 0
    while i == 0:
        choice = input("Whom would you like to delete from the database?\n(1) Individual tutor\t\t(2) All tutors\n")
        if choice == '1':
            email = input("\nEmail address of tutor to be deleted:")
            delete_Individual(email, ws)
            i = 1
        elif choice == '2':
            delete_Everyone(ws)
            i = 1
        else:
            wrongEntry()
    
    wb.save(tutor_data)
    wb.close()



def add_Individual():
    print("\n-------------Adding new tutor-------------\n")

    j = 0
    while j == 0:
        fname = input("First name:")
        lname = input("Last name:")
        email = input("Email address:")

        sql = 'INSERT INTO tutor_data VALUE("%s", "%s", "%s")' % (fname, lname, email)
        cursor.execute(sql)
        result = cursor.fetchall()
        for column in result:
            fname = column[0]
            lname = column[1]
            email = column[2]
            
        print("\n", fname, lname, email)
        i = 0
        while i == 0:
            jein = input("\nConfirm entry (1) Yes\t(0) No")
            if jein == '1':
                update_Master_List(fname, lname, email)
                db.commit()
                i = 1
            elif jein == '0':
                db.rollback()
                i = 1
            else:
                jeinError()

        i = 0
        while i == 0:
            another = input("Add another tutor? (1) Yes\t(0) No")
            if another == '1':
                i = 1
            elif another == '0':
                i = 1
                j = 1
            else:
                jeinError()



def import_Data():
    global tutor_data
    global sheet_name

    wb = load_workbook(tutor_data)
    ws = wb[sheet_name]
    total_entry = ws.max_row - 1  # -1 as the first row in the excel file is the table headings

    sql = "DELETE FROM tutor_data"
    cursor.execute(sql)
    db.commit()

    i = 0
    for i in range(total_entry):
        fname = ws['A'+str(i+2)].value
        lname = ws['B'+str(i+2)].value
        email = ws['C'+str(i+2)].value
        try:
            sql = 'INSERT INTO tutor_data VALUE("%s", "%s", "%s")' % (fname, lname, email)
            cursor.execute(sql)
            db.commit()
        except:
            continue


def edit_Particulars():
    print("\n-------------Update particulars-------------\n")
    email_tutor = input("Email address of tutor:")

    sql = "SELECT * FROM tutor_data WHERE Email = '%s'" % (email_tutor)
    cursor.execute(sql)
    result = cursor.fetchall()

    try:
        for column in result:
            fname = column[0]
            lname = column[1]

        print("\nData of", fname, lname, "will be edited.\n")
        wb = load_workbook(tutor_data)
        ws = wb[sheet_name]

        j = 0
        while j == 0:
            edit = input("What would you like to edit?\n(1) First name\t\t(3) Email address\n(2) Last Name")
            if edit == '1':
                newFname(email_tutor, ws)
                j = 1
            elif edit == '2':
                newLname(email_tutor, ws)
                j = 1
            elif edit == '3':
                newEmail(email_tutor, ws)
                j = 1
            else:
                wrongEntry()
        
        wb.save(tutor_data)
        wb.close()
        
    except:
        print("Tutor does not exist in the database.\n")




# main body
print("\n-------------Welcome to the tutor management program-------------")

i = 0
while i == 0:
    choice = input("\nChoose an option below:\n(1) Import from Excel\t\t(3) Add a tutor\n(2) Edit particulars\t\t(4) Delete Mode")
    if choice == '1':
        import_Data()
        i = 1
    elif choice == '2':
        edit_Particulars()
        i = 1
    elif choice == '3':
        add_Individual()
        i = 1
    elif choice == '4':
        delete_Mode()
        i = 1
    else:
        wrongEntry()

db.close()
