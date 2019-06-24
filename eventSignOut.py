import pymysql
import datetime
import time
import shutil
import os
from random import randint
import unidecode
from unidecode import unidecode
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

import connection
from connection import db_ip_address
from connection import db_user
from connection import db_password
from connection import db_name
from connection import event_folder
from connection import receipt_temp_folder
from connection import receipt_perm_folder
from connection import email_sender
from connection import email_sender_password
from connection import email_hiwi
from connection import server_out
from connection import port_out

db = pymysql.connect(host=db_ip_address, user=db_user, passwd=db_password, db=db_name)
cursor = db.cursor()

def random_with_N_digits(n):
    range_start = 10**(n-1)
    range_end = (10**n)-1
    return randint(range_start, range_end)


def sendShiftReceipt(fnameWL, lnameWL, email_WL, event_description, event_date):
    global email_sender
    global email_sender_password
    global server_out
    global port_out

    year, month, day = map(str, event_date.split('-'))
    event_date = day + "/" + month + "/" + year

    msg = MIMEMultipart()
    msg['From'] = email_sender
    msg['To'] = email_WL
    msg['Subject'] = "TUMi: " + event_description + " You Are Off the Waiting List"

    body = "Hi " + fnameWL + " " + lnameWL + ",\n\nWe would like to inform you that you are now off the waiting list and in the official list of participants for the event: " + event_description + " on " + event_date + ".\nDeregistration is 7 working days before the event.\n\n\nRegards,\n\nThomas Bergmann, M.A.\nOffice for Incoming Exchange Students / TUMi\n\nTechnical University of Munich\nTUM International Centre\n\nArcisstr. 21\n80333 Munich\n\nTel: + 49 89 289 25477"
    msg.attach(MIMEText(body, 'plain'))

    #msg.attach(part)
    text = msg.as_string()
    server = smtplib.SMTP(server_out, port_out)
    server.starttls()
    server.login(email_sender, email_sender_password)
    #email_receivers = [email_WL] + [email_sender] + [email_hiwi]
    email_receivers = [email_WL]
    server.sendmail(email_sender, email_receivers, text)
    server.quit()


def send_Receipt(fname, lname, event_description, receipt_temp_path, receipt_name, email_student, receipt_perm_path, waiting_list_jein):
    global email_sender
    global email_hiwi
    global email_sender_password
    global server_out
    global port_out

    msg = MIMEMultipart()
    msg['From'] = email_sender
    msg['To'] = email_student

    if waiting_list_jein == False:  # Email to be sent to ppl in the regular participants' list
        msg['Subject'] = "TUMi: " + event_description + ' Deregistration Receipt'
        body = "Hi " + fname + " " + lname + ",\n\nAttached herewith is the receipt for the event you just deregistered from.\n\n\nRegards,\nThomas Bergmann, M.A.\nOffice for Incoming Exchange Students / TUMi\n\nTechnical University of Munich\nTUM International Centre\n\nArcisstr. 21\n80333 Munich\n\nTel: + 49 89 289 25477"
    else:  # Email to be sent to ppl in the waiting list
        msg['Subject'] = "TUMi: " + event_description + ' Waiting List Deregistration Receipt'
        body = "Hi " + fname + " " + lname + ",\n\nAttached herewith is the receipt (waiting list) for the event you just deregistered from.\n\n\nRegards,\nThomas Bergmann, M.A.\nOffice for Incoming Exchange Students / TUMi\n\nTechnical University of Munich\nTUM International Centre\n\nArcisstr. 21\n80333 Munich\n\nTel: + 49 89 289 25477"


    msg.attach(MIMEText(body, 'plain'))
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(open(receipt_temp_path, 'rb').read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= " + receipt_name)
    msg.attach(part)
    text = msg.as_string()
    server = smtplib.SMTP(server_out, port_out)
    server.starttls()
    server.login(email_sender, email_sender_password)
    #email_receivers = [email_student] + [email_sender] + [email_hiwi]
    email_receivers = [email_student]
    server.sendmail(email_sender, email_receivers, text)
    server.quit()

    shutil.move(receipt_temp_path, receipt_perm_path)

    if waiting_list_jein == False:
        print("\nDeregistration receipt sent to:", email_student)
    else:
        print("\nDeregistration receipt (waiting list) sent to:", email_student)




def create_Deregistration_Receipt(event_description, event_code, event_date, fname, lname, event_price, email_student, waiting_list_jein):
    temp_folder = receipt_temp_folder + event_code  # For temp receipts, we have a subfolder for each event.
    perm_folder = receipt_perm_folder + event_code


    # Contents of receipt
    receipt_number = str(random_with_N_digits(8))
    if waiting_list_jein == False:  # For those in the regular list
        receipt_name = "Deregister_" + receipt_number + "_" + unidecode(fname + lname) + ".pdf"
        receipt_description = "Event name                             : " + event_description + " Deregistration"
    else:   # For those in the waiting list
        receipt_name = "WL_Deregister_" + receipt_number + "_" + unidecode(fname + lname) + ".pdf"
        receipt_description = "Event name                             : " + event_description + " Waiting List Deregistration"



    receipt_event_code = "Event code                              : " + event_code
    receipt_event_date = str(event_date)
    year, month, day = map(str, receipt_event_date.split('-'))
    event_date = day + "/" + month + "/" + year
    receipt_event_date = "Event date                               : " + str(event_date)
    receipt_participant_name = "Participant                               : " + fname + " " + lname
    system_timestamp = time.strftime('%d/%m/%Y  %X')
    receipt_timestamp = "Date and time of payment       :" + " " + system_timestamp
    receipt_amount = "Amount reimbursed                 : €" + event_price
    receipt_number = "Receipt number                       : " + receipt_number

    save_name = os.path.join(os.path.expanduser("~"), temp_folder, receipt_name)
    c = canvas.Canvas(save_name, pagesize=A4)
    receipt_temp_path = temp_folder + "/" + receipt_name
    receipt_perm_path = perm_folder + "/" + receipt_name
    # header text
    c.setFont(size=14, psfontname="Helvetica", leading=None)
    c.drawString(70, 670, "International Centre")
    c.drawString(70, 650, "Technical University of Munich")
    c.drawString(70, 630, "Arcisstr. 21, 80333 Munich")
    c.drawString(70, 610, "Germany")

    c.setFont(size=22, psfontname="Helvetica-Bold", leading=None)
    c.drawCentredString(x=300, y=570, text="RECEIPT")

    c.setFont(size=14, psfontname="Helvetica", leading=None)
    c.drawString(70, 530, receipt_description)
    c.drawString(70, 510, receipt_event_code)
    c.drawString(70, 490, receipt_event_date)

    c.drawString(70, 450, receipt_participant_name)
    c.drawString(70, 430, receipt_timestamp)
    c.drawString(70, 410, receipt_amount)

    c.drawString(70, 370, receipt_number)

    c.drawImage(image='index.png', x=295, y=700, preserveAspectRatio=False, width=229.2, height=79.2)
    c.showPage()
    c.save()

    send_Receipt(fname, lname, event_description, receipt_temp_path, receipt_name,
                email_student, receipt_perm_path, waiting_list_jein)  # Function created to easily on / off send receipt feature




# Main body


email_student = input("Student's email address:")
sql = "SELECT date, eventCode, description FROM events order by date ASC"
cursor.execute(sql)
db.commit()

result = cursor.fetchall()
print("\nList of all events past till future:")
time.sleep(0.8)
print("\nDate\t\t\t Code\t\t\tDescription")
for column in result:
    event_date = column[0]
    event_code = column[1]
    event_description = column[2]
    print(event_date, "\t\t", event_code, "\t\t   ", event_description)



k = 0
while k == 0:
    try:
        event_code = input("\nEvent code:")
        excel_file = event_code + '.xlsx'
        excel_path = event_folder + excel_file
        wb = load_workbook(excel_path)
        ws = wb['Sheet']
        ws_wl = wb['Waiting_List']
        k = 1  # Correct event code entered
    except:
        print("Wrong event code entered. Please try again.\n")


x = 0  # In the waiting list?
y = 0  # In the regular list?
for row_number in range(ws_wl.max_row, 1, -1):
    if ws_wl.cell(row=row_number, column=3).value == email_student:
        fname = ws_wl.cell(row=row_number, column=1).value
        lname = ws_wl.cell(row=row_number, column=2).value
        ws_wl.delete_rows(row_number)  # delete from excel
        x = 1  # Person is in the waiting list
        break   # Quit loop once the person has been found

if x == 0:  # Person not in the waiting list
    for row_number in range(ws.max_row, 1, -1):
        if ws.cell(row=row_number, column=4).value == email_student:
            fname = ws.cell(row=row_number, column=2).value
            lname = ws.cell(row=row_number, column=3).value
            ws.delete_rows(row_number)  # delete from excel
            y = 1  # Person is in the regular list
            break  # Quit loop once the person has been found


if x == 0 and y == 0:
    print("\nParticipant never registered for the event.")

elif x == 1:  # Person is in the waiting list
    sql = 'select * from events where eventCode = "%s"' % (event_code)
    cursor.execute(sql)
    result = cursor.fetchall()
    for column in result:
        event_description = column[-1]
        event_date = column[0]
        event_price = column[-2]

    # receipt
    create_Deregistration_Receipt(event_description, event_code, event_date, fname, lname, str(event_price), email_student, waiting_list_jein=True)
    sql = 'INSERT INTO event_transactions value(curdate(), curtime(), "%s", "%s", "%s", "%s", "%.2f")' % (
    event_code + " Waiting List Deregistration", email_student, fname, lname, event_price*-1)
    cursor.execute(sql)
    db.commit()

    # cash flow
    sql = 'INSERT INTO cashFlow value(now(), "%s", "%.2f")' % (event_description + " Waiting List Deregistration", event_price*-1)
    cursor.execute(sql)
    db.commit()

    # update number of participants in databank
    sql = 'UPDATE events SET inWaitingList = inWaitingList - 1 WHERE eventCode = "%s"' % (event_code)
    cursor.execute(sql)
    db.commit()

    print("Please give", fname, lname, event_price, "Euros.")

elif x == 0 and y == 1:  # Person is in the regular list
    sql = 'select * from events where eventCode = "%s"' % (event_code)
    cursor.execute(sql)
    result = cursor.fetchall()
    for column in result:
        event_description = column[-1]
        event_date = column[0]
        event_price = column[-2]

    # receipt
    create_Deregistration_Receipt(event_description, event_code, event_date, fname, lname, str(event_price), email_student, waiting_list_jein=False)
    sql = 'INSERT INTO event_transactions value(curdate(), curtime(), "%s", "%s", "%s", "%s", "%.2f")' % (
    event_code + " Deregistration", email_student, fname, lname, event_price*-1)
    cursor.execute(sql)
    db.commit()

    # cash flow
    sql = 'INSERT INTO cashFlow value(now(), "%s", "%.2f")' % (event_description + " Deregistration", event_price*-1)
    cursor.execute(sql)
    db.commit()

    # update number of participants in databank
    sql = 'UPDATE events SET numberOfParticipants = numberOfParticipants - 1 WHERE eventCode = "%s"' % (event_code)
    cursor.execute(sql)
    db.commit()

    print("Please give", fname, lname, event_price, "Euros.")


sql = 'SELECT * from events WHERE eventCode = "%s"' % (event_code)
cursor.execute(sql)
result = cursor.fetchall()
for column in result:
    event_date = column[0]
    qty_participants = column[2]
    max_participants = column[3]
    qty_in_wl = column[4]
    event_description = column[-1]
if qty_participants == (max_participants-1) and qty_in_wl >= 1:
    sql1 = 'UPDATE events SET numberOfParticipants = numberOfParticipants + 1 WHERE eventCode = "%s"' % (event_code)
    cursor.execute(sql1)
    sql2 = 'UPDATE events SET inWaitingList = inWaitingList - 1 WHERE eventCode = "%s"' % (event_code)
    cursor.execute(sql2)
    db.commit()

    if event_date > datetime.date.today():
        fname_wl = ws_wl.cell(row=2, column=1).value
        lname_wl = ws_wl.cell(row=2, column=2).value
        email_wl = ws_wl.cell(row=2, column=3).value
        country_wl = ws_wl.cell(row=2, column=4).value
        department_wl = ws_wl.cell(row=2, column=5).value
        status_wl = ws_wl.cell(row=2, column=6).value
        ws_wl.delete_rows(2)
        ws.append(["", fname_wl, lname_wl, email_wl, country_wl, department_wl, status_wl])

        sendShiftReceipt(fname_wl, lname_wl, email_wl, event_description, str(event_date))

wb.save(excel_path)
wb.close()

db.close()