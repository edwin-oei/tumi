import pymysql
import datetime
import time
import unidecode
from unidecode import unidecode
import os
import shutil
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import atexit
import reportlab

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4

from random import randint

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

import connection
from connection import db_ip_address
from connection import db_user
from connection import db_password
from connection import db_name
from connection import student_data
from connection import event_folder
from connection import receipt_temp_folder
from connection import receipt_perm_folder
from connection import email_sender
from connection import email_sender_password
from connection import email_hiwi
from connection import server_out
from connection import port_out


# Connect immediately to databank
db = pymysql.connect(host=db_ip_address, user=db_user, passwd=db_password, db=db_name)  # Establish connection to databank
cursor = db.cursor()

original_qty_into_regular_list = 0
original_qty_into_wl = 0

def random_with_N_digits(n):
    range_start = 10**(n-1)
    range_end = (10**n)-1
    return randint(range_start, range_end)



def jeinError():
    print("Please enter either 1 or 0.\n")



def check_participant_DB(email_students, qty_students):
    global student_data

    sql = [""] * qty_students
    for i in range(qty_students):
        sql[i] = "SELECT * FROM student_data WHERE Email = '%s'" % (email_students[i])
        cursor.execute(sql[i])
        results = cursor.fetchall()
        for column in results:
            fname = column[0]
            lname = column[1]

        try:
            print(fname, lname, "is in the databank.")
            del fname  # Erase contents of placeholder. Ensure error comes up if fname is empty after being filled previously
            del lname
        except:
            print("\nParticipant with the email address", email_students[i], "is not in the databank.")
            fname_new = input("First name:")
            lname_new = input("Last name:")
            country_new = input("Country:")
            department_new = input("Department:")
            j = 0
            while j == 0:
                status_new = input("E / D:")
                if status_new == 'E' or status_new == 'e':
                    j = 1
                elif status_new == 'D' or status_new == 'd':
                    j = 1
                else:
                    print("Please enter either E or D.")

            sql2 = 'INSERT INTO student_data value("%s", "%s", "%s", "%s", "%s", "%s")' % (fname_new, lname_new, email_students[i], country_new, department_new, status_new)
            cursor.execute(sql2)
            db.commit()

            wb = load_workbook(student_data)  # Update excel list of participants / erasmi
            ws = wb['Sheet1']
            ws.append([fname_new, lname_new, email_students[i], country_new, department_new, status_new])
            wb.save(student_data)


def cashFlow(event_description, price_perpax, waiting_list_jein):
    if waiting_list_jein == False:
        sql = 'INSERT INTO cashFlow VALUE(now(), "%s", "%.2f")' % (event_description + " Registration", price_perpax)
    else:
        sql = 'INSERT INTO cashFlow VALUE(now(), "%s", "%.2f")' % (event_description + " Registration - Waiting List", price_perpax)
    cursor.execute(sql)
    db.commit()



def send_Receipt(fname, lname, event_description, receipt_temp_path, receipt_name, email_student, receipt_perm_path, waiting_list_jein):
    global email_sender
    global email_hiwi
    global email_sender_password
    global server_out
    global port_out

    msg = MIMEMultipart()
    msg['From'] = email_sender
    msg['To'] = email_student

    if waiting_list_jein == False:  # Email to be sent to ppl in the regular participants' list
        msg['Subject'] = "TUMi: " + event_description + ' Registration Receipt'
        body = "Hi " + fname + " " + lname + ",\n\nAttached herewith is the receipt for the event you just signed up for.\n\n\nRegards,\nThomas Bergmann, M.A.\nOffice for Incoming Exchange Students / TUMi\n\nTechnical University of Munich\nTUM International Centre\n\nArcisstr. 21\n80333 Munich\n\nTel: + 49 89 289 25477"
    else:  # Email to be sent to ppl in the waiting list
        msg['Subject'] = "TUMi: " + event_description + ' Waiting List Registration Receipt'
        body = "Hi " + fname + " " + lname + ",\n\nAttached herewith is the receipt (waiting list) for the event you just signed up for.\n\n\nRegards,\nThomas Bergmann, M.A.\nOffice for Incoming Exchange Students / TUMi\n\nTechnical University of Munich\nTUM International Centre\n\nArcisstr. 21\n80333 Munich\n\nTel: + 49 89 289 25477"


    msg.attach(MIMEText(body, 'plain'))
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(open(receipt_temp_path, 'rb').read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= " + receipt_name)
    msg.attach(part)
    text = msg.as_string()
    server = smtplib.SMTP(server_out, port_out)
    server.starttls()
    server.login(email_sender, email_sender_password)
    #email_receivers = [email_student] + [email_sender] + [email_hiwi]
    email_receivers = [email_student]
    server.sendmail(email_sender, email_receivers, text)
    server.quit()

    shutil.move(receipt_temp_path, receipt_perm_path)

    if waiting_list_jein == False:
        print("\ne-receipt sent to:", email_student)
    else:
        print("\ne-receipt (waiting list) sent to:", email_student)



def create_Receipt(event_description, event_code, event_date, fname, lname, price_perpax, email_student, waiting_list_jein):
    temp_folder = receipt_temp_folder + event_code  # For temp receipts, we have a subfolder for each event.
    perm_folder = receipt_perm_folder + event_code

    if not os.path.exists(temp_folder):  # Create receipt folder for each event if not done yet
        os.makedirs(temp_folder)
    if not os.path.exists(perm_folder):  # Create receipt folder for each event if not done yet
        os.makedirs(perm_folder)

    # Contents of receipt
    receipt_number = str(random_with_N_digits(8))
    if waiting_list_jein == False:  # For those in the regular list
        receipt_name = "Register_" + receipt_number + "_" + unidecode(fname + lname) + ".pdf"
        receipt_description = "Event name                             : " + event_description
    else:   # For those in the waiting list
        receipt_name = "WL_" + receipt_number + "_" + unidecode(fname + lname) + ".pdf"
        receipt_description = "Event name                             : " + event_description + " (Waiting List)"


    receipt_event_code = "Event code                              : " + event_code
    receipt_event_date = str(event_date)
    year, month, day = map(str, receipt_event_date.split('-'))
    event_date = day + "/" + month + "/" + year
    receipt_event_date = "Event date                               : " + str(event_date)
    receipt_participant_name = "Participant                               : " + fname + " " + lname
    system_timestamp = time.strftime('%d/%m/%Y  %X')
    receipt_timestamp = "Date and time of payment       :" + " " + system_timestamp
    receipt_amount = "Amount paid                            : €" + price_perpax
    receipt_number = "Receipt number                       : " + receipt_number

    save_name = os.path.join(os.path.expanduser("~"), temp_folder, receipt_name)
    c = canvas.Canvas(save_name, pagesize=A4)
    receipt_temp_path = temp_folder + "/" + receipt_name
    receipt_perm_path = perm_folder + "/" + receipt_name
    # header text
    c.setFont(size=14, psfontname="Helvetica", leading=None)
    c.drawString(70, 670, "International Centre")
    c.drawString(70, 650, "Technical University of Munich")
    c.drawString(70, 630, "Arcisstr. 21, 80333 Munich")
    c.drawString(70, 610, "Germany")

    c.setFont(size=22, psfontname="Helvetica-Bold", leading=None)
    c.drawCentredString(x=300, y=570, text="RECEIPT")

    c.setFont(size=14, psfontname="Helvetica", leading=None)
    c.drawString(70, 530, receipt_description)
    c.drawString(70, 510, receipt_event_code)
    c.drawString(70, 490, receipt_event_date)

    c.drawString(70, 450, receipt_participant_name)
    c.drawString(70, 430, receipt_timestamp)
    c.drawString(70, 410, receipt_amount)

    c.drawString(70, 370, receipt_number)

    c.drawString(70, 310, "*Please bring this receipt to deregister from the event.")
    c.drawString(70, 290, "**Deregistration deadline: 7 days before the event.")

    c.drawImage(image='index.png', x=295, y=700, preserveAspectRatio=False, width=229.2, height=79.2)
    c.showPage()
    c.save()

    send_Receipt(fname, lname, event_description, receipt_temp_path, receipt_name,
                email_student, receipt_perm_path, waiting_list_jein)  # Function created to easily on / off send receipt feature




def confirm_Register(email_students, qty_students, event_code, waiting_list_jein):
    global event_folder
    global sent_to_everyone
    global k_sent_regular_list
    global k_sent_waiting_list
    global cancel_registration
    global original_qty_into_regular_list
    global original_qty_into_wl
    
    sent_to_everyone = 0  # Has to be set to 0 again after regular list is done

    excelFile = event_folder + event_code + ".xlsx"

    sql = 'SELECT * FROM events WHERE eventCode = "%s"' % (event_code)
    cursor.execute(sql)
    result = cursor.fetchall()
    for column in result:
        price_perpax = column[-2]  # For receipt creation where the price for 1 person must be stated
        event_description = column[-1]  # For receipt
        event_date = column[0]  # For receipt

    if qty_students == 1:
        print("\nPlease ask him/her to pay €", price_perpax)
        jein = input("\nHas he/she paid for the event? (1) Yes\t(0) No")
    elif qty_students > 1:
        print("\nPlease ask them to pay €", price_perpax*qty_students)
        jein = input("\nHave they paid for the event? (1) Yes\t(0) No")
        
    i = 0
    while i == 0:
        if jein == '1':
            i = 1
            cancel_registration = 0  # Registration confirmed
            for i in range(qty_students):
                sql1 = 'SELECT * FROM student_data WHERE Email = "%s"' % (email_students[i])
                cursor.execute(sql1)
                result = cursor.fetchall()
                for column in result:
                    fname = column[0]
                    lname = column[1]
                    country = column[3]
                    department = column[4]
                    status = column[5]

                create_Receipt(event_description, event_code, event_date, fname, lname, str(price_perpax), email_students[i], waiting_list_jein)
                cashFlow(event_description, price_perpax, waiting_list_jein)

                sql2 = 'INSERT INTO event_transactions value(curdate(), curtime(), "%s", "%s", "%s", "%s", "%.2f")' % (
                    event_code, email_students[i], fname, lname, price_perpax)
                cursor.execute(sql2)
                db.commit()

                wb = load_workbook(excelFile)  # Open workbook and appends it in for loop so as to allow for parallel work.
                if waiting_list_jein == False:
                    ws = wb['Sheet']
                    ws.append(["", fname, lname, email_students[i], country, department, status])
                    k_sent_regular_list = i + 1
                else:
                    ws = wb['Waiting_List']
                    ws.append([fname, lname, email_students[i], country, department, status])
                    k_sent_waiting_list = i + 1

                wb.save(excelFile)
                wb.close()

            sent_to_everyone = 1  # Managed to send the receipt to everyone.

        elif jein == '0':
            if waiting_list_jein == False:
                sql1 = 'UPDATE events SET numberOfParticipants = numberOfParticipants - "%i" WHERE eventCode = "%s"' % (qty_students, event_code)
                sql2 = 'UPDATE events SET inWaitingList = inWaitingList - "%i" WHERE eventCode = "%s"' % (original_qty_into_wl, event_code)
                cursor.execute(sql1)
                cursor.execute(sql2)
            else:
                sql1 = 'UPDATE events SET inWaitingList = inWaitingList - "%i" WHERE eventCode = "%s"' % (qty_students, event_code)
                cursor.execute(sql1)

            db.commit()
            cancel_registration = 1  # Registration cancelled
            print("Registration cancelled. Program stopped.")
            break

        else:
            jeinError()  # If anything other than 1 or 0 is entered. Repeat loop





def enter_Participant_Emails(qty_students, event_code, waiting_list_jein):
    email_students = [""] * qty_students
    for i in range(qty_students):  # Get email addresses of the participants
        if waiting_list_jein == False:
            print("Email address of participant" + " " + str(i + 1) + ":")
        else:
            print("Email address of participant" + " " + str(i + 1) + " in the waiting list:")    
        email_students[i] = input()

    check_participant_DB(email_students, qty_students)  # Check if they are already in the databank
    confirm_Register(email_students, qty_students, event_code, waiting_list_jein)




def check_Availability(event_code, qty_participants):
    global original_qty_into_regular_list
    global original_qty_into_wl
    global cancel_registration

    sql = 'SELECT * FROM events WHERE eventCode = "%s"' % (event_code)
    cursor.execute(sql)
    result = cursor.fetchall()
    for column in result:
        free_spots = column[3] - column[2]
        free_wl = column[5] - column[4]

    if qty_participants > free_spots and free_wl > 0:  # Case 1: Not enough spots available, waiting list spots available but not necessarily enough
        print("There are only", free_spots, "spots available.\n")  # Case 1: Available spots in waiting list will change for sure
        j = 0
        while j == 0:
            if (qty_participants - free_spots) > free_wl:  # Case 1a: too many interested, such that some can't even be in the waiting list
                qty_into_wl = free_wl# Number of people going into the waiting list = number of spots available in the waiting list
                print("Enter", qty_into_wl, "people into the waiting list and leave out",
                      qty_participants - free_spots - free_wl, "people?")

            elif (qty_participants - free_spots) <= free_wl:  # Case 1b: Available waiting list spots is sufficient
                qty_into_wl = (qty_participants - free_spots)  # Number of people going into the waiting list = difference
                print("Enter", qty_into_wl, "people into the waiting list?")

            original_qty_into_wl = qty_into_wl  # For global settings in case email cant be sent
            original_qty_into_regular_list = free_spots

            jein = input("(1) Yes\t(0) No")
            if jein == '1':
                j = 1  # entered 1, get out of loop
                sql1 = 'UPDATE events SET inWaitingList = inWaitingList + "%i" WHERE eventCode = "%s"' % (qty_into_wl, event_code)
                cursor.execute(sql1)
                db.commit()

                if free_spots > 0:  # Participants' list becomes full. Number of participants increase by the number of available spots left
                    sql2 = 'UPDATE events SET numberOfParticipants = numberOfParticipants + "%i" WHERE eventCode = "%s"' % (free_spots, event_code)  # update participants' list
                    cursor.execute(sql2)
                    db.commit()
                    enter_Participant_Emails(free_spots, event_code, waiting_list_jein=False)
                    if cancel_registration == 0:
                        enter_Participant_Emails(qty_into_wl, event_code, waiting_list_jein=True)

                elif free_spots <= 0:  # participants' list was already full.
                    enter_Participant_Emails(qty_into_wl, event_code, waiting_list_jein=True)

            elif jein == '0':  # no changes made to databank
                j = 1  # entered 0. get out of loop

            else:
                jeinError()  # Entered something other than 1 or 0. Repeat loop


    elif qty_participants <= free_spots:  # Case 2: There are more spots available than people wanting to register for the event
        original_qty_into_regular_list = qty_participants
        sql1 = 'UPDATE events SET numberOfParticipants = numberOfParticipants + "%i" WHERE eventCode = "%s"' % (qty_participants, event_code)
        cursor.execute(sql1)
        db.commit()
        enter_Participant_Emails(qty_participants, event_code, waiting_list_jein=False)
        

    elif free_spots == 0 and free_wl == 0:  # Case 3: Event fully booked. Waiting list already full
        print("Event fully booked and waiting list full.\n")



def check_Event():  # Prints list of all events taking place from today till the future
    sql = "SELECT * FROM events WHERE date BETWEEN curdate() AND '2099-12-31' ORDER BY date"
    cursor.execute(sql)

    result = cursor.fetchall()
    print("\nDate\t\t\tCode\t\tPlaces left/total\t\tWaiting list spots left/total\t\tPrice\t\tDescription")
    for column in result:  # Unpack info
        event_date = column[0]
        event_code = column[1]
        num_participants = column[2]
        max_spots = column[3]
        num_wl = column[4]
        max_wl = column[5]
        event_price = column[6]
        event_description = column[7]

        print(event_date, "\t\t", event_code, "\t\t   ", max_spots-num_participants, " /", max_spots, "\t\t\t\t\t", max_wl-num_wl, " /", max_wl, "\t\t\t\t\t", event_price,
              "\t\t  ", event_description)




# Main Body


i = 0
while i == 0:
    try:  # Ensure that event code entered is correct, otherwise try again
        check_Event()  # Prints all created events taking place from today till the future
        event_code = input("Event code:")
        codeForReset = event_code
        sql = "SELECT * FROM events WHERE eventCode = '%s'" % (event_code)
        cursor.execute(sql)
        result = cursor.fetchall()
        for column in result:
            event_date = column[0]
            event_description = column[7]
        print(event_description, "on", event_date, "selected.\n")
        i = 1  # Correct event code given, ie event code exists in the databank
    except:
        print("Wrong event code. Please try again.")

i = 0
cancel_registration = 1  # Control variable for registration confirmation
sent_to_everyone = 0  # Control variable for sending of ALL receipts
k_sent_regular_list = 0  # Control variable for number of receipts successfully sent to ppl in the regular list
k_sent_waiting_list = 0  # Control variable for number of receipts successfully sent to ppl in the waiting list
cancel_registration = 0

try:
    while i == 0:
        try:
            qty_participants = int(input("Number of interested participants:"))
            i = 1  # Integer entered, get out of loop.
        except:
            print("Please enter an integer")

    check_Availability(event_code, qty_participants)

except:
    if sent_to_everyone == 0:  # email not sent to everyone
        sql1 = 'UPDATE events SET numberOfParticipants = numberOfParticipants - "%i" WHERE eventCode = "%s"' % (original_qty_into_regular_list - k_sent_regular_list, event_code)
        sql2 = 'UPDATE events SET inWaitingList = inWaitingList - "%i" WHERE eventCode = "%s"' % (original_qty_into_wl - k_sent_waiting_list, event_code)
        cursor.execute(sql1)
        cursor.execute(sql2)
        db.commit()
        print("\nFailed to send e-receipt. Please check internet connection or validity of the email address entered.")


db.close()  # Close connection to databank