import pymysql
import datetime
import time
import os
import pandas as pd
from pandas import ExcelWriter
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

import connection
from connection import db_ip_address
from connection import db_user
from connection import db_password
from connection import db_name
from connection import student_data

db = pymysql.connect(host=db_ip_address, user=db_user, passwd=db_password, db=db_name)
cursor = db.cursor()

sheet_name = 'Sheet1'

def jeinError():
    print("Please enter either 1 or 0.\n")



def wrongEntry():
    print("Wrong entry. Please try again.\n")



def update_Master_List(fname, lname, email, country, department, status):
    global student_data
    global sheet_name
    wb = load_workbook(student_data)
    ws = wb[sheet_name]
    ws.append([fname, lname, email, country, department, status])
    wb.save(student_data)
    wb.close()




def updatedEntry(result):

    for column in result:
        fname = column[0]
        lname = column[1]
        email = column[2]
        country = column[3]
        department = column[4]
        status = column[5]
    print(fname, lname, email, country, department, status)
    i = 0
    while i == 0:
        jein = input("Confirm entry (1) Yes\t(0) No")
        if jein == '1':
            db.commit()
            i = 1
            return 1
        elif jein == '0':
            db.rollback()
            i = 1
        else:
            jeinError()


def newFname(email, ws):

    new_fname = input("New first name:")
    sql = 'UPDATE student_data SET First_Name = "%s" WHERE Email = "%s"' % (new_fname, email)
    cursor.execute(sql)
    sql1 = "SELECT * FROM student_data WHERE Email = '%s'" % (email)
    cursor.execute(sql1)
    result = cursor.fetchall()
    if updatedEntry(result) == 1:
        for rowNum in range(ws.max_row, 1, -1):
            if ws.cell(row=rowNum, column=3).value == email:
                ws.cell(row=rowNum, column=1).value = new_fname


def newLname(email, ws):

    new_lname = input("New last name:")
    sql = 'UPDATE student_data SET Last_Name = "%s" WHERE Email = "%s"' % (new_lname, email)
    cursor.execute(sql)
    sql1 = "SELECT * FROM student_data WHERE Email = '%s'" % (email)
    cursor.execute(sql1)
    result = cursor.fetchall()
    if updatedEntry(result) == 1:
        for rowNum in range(ws.max_row, 1, -1):
            if ws.cell(row=rowNum, column=3).value == email:
                ws.cell(row=rowNum, column=2).value = new_lname



def newCountry(email, ws):

    new_country = input("New country:")
    sql = "UPDATE student_data SET Country = '%s' WHERE Email = '%s'" % (new_country, email)
    cursor.execute(sql)
    sql1 = "SELECT * FROM student_data WHERE Email = '%s'" % (email)
    cursor.execute(sql1)
    result = cursor.fetchall()
    if updatedEntry(result) == 1:
        for rowNum in range(ws.max_row, 1, -1):
            if ws.cell(row=rowNum, column=3).value == email:
                ws.cell(row=rowNum, column=4).value = new_country


def newDepartment(email, ws):

    new_department = input("New department:")
    sql = "UPDATE student_data SET Department = '%s' WHERE Email = '%s'" % (new_department, email)
    cursor.execute(sql)
    sql1 = "SELECT * FROM student_data WHERE Email = '%s'" % (email)
    cursor.execute(sql1)
    result = cursor.fetchall()
    if updatedEntry(result) == 1:
        for rowNum in range(ws.max_row, 1, -1):
            if ws.cell(row=rowNum, column=3).value == email:
                ws.cell(row=rowNum, column=5).value = new_department
        


def newStatus(email, ws):
    i = 0
    while i == 0:
        new_status = input("New E/D:")
        if new_status == 'E' or new_status == 'e':
            i = 1
        elif new_status == 'D' or new_status == 'd':
            i = 1
        else:
            print("Please enter either E or D.\n")

    sql = "UPDATE student_data SET ED = '%s' WHERE Email = '%s'" % (new_status, email)
    cursor.execute(sql)
    sql1 = "SELECT * FROM student_data WHERE Email = '%s'" % (email)
    cursor.execute(sql1)
    result = cursor.fetchall()
    if updatedEntry(result) == 1:
        for rowNum in range(ws.max_row, 1, -1):
            if ws.cell(row=rowNum, column=3).value == email:
                ws.cell(row=rowNum, column=6).value = new_status
        


def newEmail(email, ws):

    new_email = input("New email address:")
    sql = "UPDATE student_data SET Email = '%s' WHERE Email = '%s'" % (new_email, email)
    cursor.execute(sql)
    sql1 = "SELECT * FROM student_data WHERE Email = '%s'" % (new_email)
    cursor.execute(sql1)
    result = cursor.fetchall()
    if updatedEntry(result) == 1:
        for rowNum in range(ws.max_row, 1, -1):
            if ws.cell(row=rowNum, column=3).value == email:
                ws.cell(row=rowNum, column=3).value = new_email


def delete_Erasmi(ws):

    sql = "DELETE FROM student_data WHERE ED = '%s'" % ('E')
    cursor.execute(sql)

    i = 0
    while i == 0:
        jein = input("Delete all Erasmi. Are you sure? (1) Yes\t(0) No")
        if jein == '1':
            j = 0
            while j == 0:
                    jein2 = input("ARE YOU FUCKING SURE??? (1) Yes\t(0) No")
                    if jein2 == '1':
                        for rowNum in range(ws.max_row, 1, -1):
                            if ws.cell(row=rowNum, column=6).value == 'E':
                                ws.delete_rows(rowNum)
                        db.commit()
                        i = 1
                        j = 1
                    elif jein2 == '0':
                        db.rollback()
                        i = 1
                        j = 1
                    else:
                        jeinError()
        elif jein == '0':
            db.rollback()
            i = 1
        else:
            jeinError()


def delete_Degree(ws):

    sql = "DELETE FROM student_data WHERE ED = '%s'" % ('D')
    cursor.execute(sql)

    i = 0
    while i == 0:
        jein = input("Delete all degree students. Are you sure? (1) Yes\t(0) No")
        if jein == '1':
            j = 0
            while j == 0:
                    jein2 = input("ARE YOU FUCKING SURE??? (1) Yes\t(0) No")
                    if jein2 == '1':
                        for rowNum in range(ws.max_row, 1, -1):
                            if ws.cell(row=rowNum, column=6).value == 'D':
                                ws.delete_rows(rowNum)
                        db.commit()
                        i = 1
                        j = 1
                    elif jein2 == '0':
                        db.rollback()
                        i = 1
                        j = 1
                    else:
                        jeinError()
        elif jein == '0':
            db.rollback()
            i = 1
        else:
            jeinError()


def delete_Everyone(ws):

    sql = "DELETE FROM student_data"
    cursor.execute(sql)

    i = 0
    while i == 0:
        jein = input("Delete everyone. Are you sure? (1) Yes\t(0) No")
        if jein == '1':
            j = 0
            while j == 0:
                    jein2 = input("ARE YOU FUCKING SURE??? (1) Yes\t(0) No")
                    if jein2 == '1':
                        db.commit()
                        ws.delete_rows(2, ws.max_row)
                        i = 1
                        j = 1
                    elif jein2 == '0':
                        db.rollback()
                        i = 1
                        j = 1
                    else:
                        jeinError()
        elif jein == '0':
            db.rollback()
            i = 1
        else:
            jeinError()


def delete_Individual(email, ws):

    sql = "SELECT * FROM student_data WHERE Email = '%s'" % (email)
    cursor.execute(sql)
    result = cursor.fetchall()
    for column in result:
        fname = column[0]
        lname = column[1]

    try:
        print("\nData of", fname, lname, "will be deleted.")

        i = 0
        while i == 0:
            jein = input("Confirm? (1) Yes\t(0) No")
            if jein == '1':
                i = 1
                sql1 = "DELETE FROM student_data WHERE Email = '%s'" % (email)
                cursor.execute(sql1)
                db.commit()
                for rowNum in range(ws.max_row, 1, -1):
                    if ws.cell(row=rowNum, column=3).value == email:
                        ws.delete_rows(rowNum)
            elif jein == '0':
                db.rollback()
                i = 1
            else:
                jeinError()
    except:
        print("Person does not exist in the database to begin with.\n")



def delete_Mode():

    print("\n-------------Delete mode-------------\n")
    
    wb = load_workbook(student_data)
    ws = wb[sheet_name]
    i = 0
    while i == 0:
        choice = input("What would you like to delete from the table?\n(1) All exchange students\t\t(3) All degree Students\n(2) Individual\t\t\t\t\t(4) Everyone")
        if choice == '1':
            delete_Erasmi(ws)
            i = 1
        elif choice == '2':
            email = input("\nEmail address of student to be deleted:")
            delete_Individual(email, ws)
            i = 1
        elif choice == '3':
            delete_Degree(ws)
            i = 1
        elif choice == '4':
            delete_Everyone(ws)
            i = 1
        else:
            wrongEntry()
    
    wb.save(student_data)
    wb.close()



def add_Individual():
    print("\n-------------Adding new student-------------\n")

    j = 0
    while j == 0:
        fname = input("First name:")
        lname = input("Last name:")
        email = input("Email address:")
        country = input("Country:")
        department = input("Department at TUM:")
        i = 0
        while i == 0:
            status = input("E or D Student?")
            if status == 'D' or status == 'd':
                i = 1
            elif status == 'E' or status == 'e':
                i = 1
            else:
                print("Please enter either E or D.\n")

        sql = 'INSERT INTO student_data VALUE("%s", "%s", "%s", "%s", "%s", "%s")' % (fname, lname, email, country, department, status)
        cursor.execute(sql)
        result = cursor.fetchall()
        for column in result:
            fname = column[0]
            lname = column[1]
            email = column[2]
            country = column[3]
            department = column[4]
            status = column[5]
        print(fname, lname, email, country, department, status)
        i = 0
        while i == 0:
            jein = input("\nConfirm entry (1) Yes\t(0) No")
            if jein == '1':
                update_Master_List(fname, lname, email, country, department, status)
                db.commit()
                i = 1
            elif jein == '0':
                db.rollback()
                i = 1
            else:
                jeinError()

        i = 0
        while i == 0:
            another = input("Add another person? (1) Yes\t(0) No")
            if another == '1':
                i = 1
            elif another == '0':
                i = 1
                j = 1
            else:
                jeinError()



def import_Data():
    global student_data
    global sheet_name

    wb = load_workbook(student_data)
    ws = wb[sheet_name]
    total_entry = ws.max_row - 1  # -1 as the first row in the excel file is the table headings

    sql = "DELETE FROM student_data"
    cursor.execute(sql)
    db.commit()

    i = 0
    for i in range(total_entry):
        fname = ws['A'+str(i+2)].value
        lname = ws['B'+str(i+2)].value
        email = ws['C'+str(i+2)].value
        country = ws['D'+str(i+2)].value
        department = ws['E'+str(i+2)].value
        status = ws['F'+str(i+2)].value
        try:
            sql = 'INSERT INTO student_data VALUE("%s", "%s", "%s", "%s", "%s", "%s")' % (fname, lname, email, country, department, status)
            cursor.execute(sql)
            db.commit()
        except:
            continue


def edit_Particulars():
    print("\n-------------Update particulars-------------\n")
    email_student = input("Email address of student:")

    sql = "SELECT * FROM student_data WHERE Email = '%s'" % (email_student)
    cursor.execute(sql)
    result = cursor.fetchall()

    try:
        for column in result:
            fname = column[0]
            lname = column[1]

        print("\nData of", fname, lname, "will be edited.")
        wb = load_workbook(student_data)
        ws = wb[sheet_name]

        j = 0
        while j == 0:
            edit = input("What would you like to edit?\n(1) First name\t\t\t(4) Country\n(2) Last name\t\t\t(5) Department\n(3) Email\t\t\t\t(6) E/D")
            if edit == '1':
                newFname(email_student, ws)
            elif edit == '2':
                newLname(email_student, ws)
            elif edit == '3':
                newEmail(email_student, ws)
            elif edit == '4':
                newCountry(email_student, ws)
            elif edit == '5':
                newDepartment(email_student, ws)
            elif edit == '6':
                newStatus(email_student, ws)
            else:
                wrongEntry()

            i = 0
            while i == 0:
                jein = input("Edit other particulars of the same person? (1) Yes\t(0) No")
                if jein == '1':
                    i = 1
                elif jein == '0':
                    i = 1
                    j = 1
                else:
                    jeinError()
        
        wb.save(student_data)
        wb.close()
        
    except:
        print("Person does not exist in the databank.\n")




# main body
print("\n-------------Welcome to the student management program-------------")

i = 0
while i == 0:
    choice = input("\nChoose an option below:\n(1) Import from Excel\t\t(3) Add an individual\n(2) Edit particulars\t\t(4) Delete Mode")
    if choice == '1':
        import_Data()
        i = 1
    elif choice == '2':
        edit_Particulars()
        i = 1
    elif choice == '3':
        add_Individual()
        i = 1
    elif choice == '4':
        delete_Mode()
        i = 1
    else:
        wrongEntry()

db.close()
